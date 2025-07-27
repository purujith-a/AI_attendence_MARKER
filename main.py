import cv2
import os
from deepface import DeepFace
from openpyxl import load_workbook

# Load Excel file
wb = load_workbook("attendance.xlsx")
ws = wb.active

# Get registered students from Excel
registered_names = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1)]

# Load group photo
group_img_path = "group.jpg"
group_img = cv2.imread(group_img_path)
if group_img is None:
    print("❌ ERROR: group.jpg not found or unreadable.")
    exit()
else:
    print("[INFO] Group photo loaded successfully.")

# Detect faces in the group image
detected_faces = DeepFace.extract_faces(img_path=group_img_path, enforce_detection=False)

# Load individual student images
known_faces_dir = "student_images/"
recognized_names = set()

for face in detected_faces:
    x = face['facial_area']['x']
    y = face['facial_area']['y']
    w = face['facial_area']['w']
    h = face['facial_area']['h']
    face_img = group_img[y:y+h, x:x+w]

    # Temporarily save the cropped face
    temp_path = "temp_face.jpg"
    cv2.imwrite(temp_path, face_img)

    for student_file in os.listdir(known_faces_dir):
        student_name = os.path.splitext(student_file)[0]
        student_img_path = os.path.join(known_faces_dir, student_file)

        try:
            result = DeepFace.verify(img1_path=student_img_path, img2_path=temp_path, enforce_detection=False)
            if result["verified"]:
                recognized_names.add(student_name)
                print(f"[✔] {student_name} marked present")
                break
        except:
            pass

    os.remove(temp_path)

# Mark attendance in Excel
for row in ws.iter_rows(min_row=2, max_col=2):
    name_cell, status_cell = row
    if name_cell.value in recognized_names:
        status_cell.value = "Present"
    else:
        status_cell.value = "Absent"

wb.save("attendance.xlsx")
print("[✅] Attendance successfully updated in Excel!")